from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import ConfiguracaoSistema, ProjetoPlantao
from app.utils.formatters import formatar_data_br


class ExportacaoErro(Exception):
    pass


class ExcelExportService:
    CABECALHO = [
        "Responsável",
        "Mês",
        "Dia Semana",
        "Início",
        "Fim",
        "Data",
        "Tipo",
        "Total Horas",
        "Valor",
        "Nº Chamado",
        "Solicitante",
        "Cliente",
        "Nível",
        "Observação",
    ]

    def export(
        self, projeto: ProjetoPlantao, config: ConfiguracaoSistema, output_path: str | Path
    ) -> Path:
        if not projeto.lancamentos:
            raise ExportacaoErro("Não há lançamentos para exportar.")

        caminho = Path(output_path)
        caminho.parent.mkdir(parents=True, exist_ok=True)

        try:
            workbook = self._build_workbook(projeto, config)
            workbook.save(caminho)
            return caminho
        except PermissionError as exc:
            raise ExportacaoErro(
                "Não foi possível salvar o Excel. Feche o arquivo se ele estiver aberto."
            ) from exc
        except OSError as exc:
            raise ExportacaoErro(f"Falha ao exportar o Excel: {exc}") from exc

    def _build_workbook(
        self, projeto: ProjetoPlantao, config: ConfiguracaoSistema
    ) -> Workbook:
        if config.template_excel and Path(config.template_excel).exists():
            workbook = load_workbook(config.template_excel)
        else:
            workbook = Workbook()

        for nome in ["Lançamentos", "Resumo Responsável", "Resumo Mês Tipo"]:
            if nome in workbook.sheetnames:
                del workbook[nome]

        if workbook.sheetnames == ["Sheet"]:
            del workbook["Sheet"]

        ws_lancamentos = workbook.create_sheet("Lançamentos")
        ws_responsavel = workbook.create_sheet("Resumo Responsável")
        ws_mes_tipo = workbook.create_sheet("Resumo Mês Tipo")

        self._fill_launches_sheet(ws_lancamentos, projeto)
        self._fill_summary_by_responsible(ws_responsavel, projeto)
        self._fill_summary_by_month_type(ws_mes_tipo, projeto)
        return workbook

    def _fill_launches_sheet(self, ws, projeto: ProjetoPlantao) -> None:
        ws.append(self.CABECALHO)
        for lancamento in projeto.lancamentos:
            ws.append(
                [
                    lancamento.responsavel,
                    lancamento.mes,
                    lancamento.dia_semana,
                    lancamento.inicio,
                    lancamento.fim,
                    formatar_data_br(lancamento.data),
                    lancamento.tipo,
                    lancamento.total_horas,
                    lancamento.valor,
                    lancamento.numero_chamado,
                    lancamento.solicitante,
                    lancamento.cliente,
                    lancamento.nivel,
                    lancamento.observacao,
                ]
            )
        self._stylize_table(ws, currency_columns={9}, hours_columns={8})

    def _fill_summary_by_responsible(self, ws, projeto: ProjetoPlantao) -> None:
        header = [
            "Responsável",
            "Horas Sobre aviso",
            "Valor Sobre aviso",
            "Horas Suporte",
            "Valor Suporte",
            "Horas Totais",
            "Valor Total",
        ]
        ws.append(header)

        agrupado: dict[str, dict] = {}
        for item in projeto.lancamentos:
            resumo = agrupado.setdefault(
                item.responsavel,
                {
                    "sobre_horas": 0.0,
                    "sobre_valor": 0.0,
                    "suporte_horas": 0.0,
                    "suporte_valor": 0.0,
                },
            )
            if item.tipo == "Sobre aviso":
                resumo["sobre_horas"] += item.total_horas
                resumo["sobre_valor"] += item.valor
            else:
                resumo["suporte_horas"] += item.total_horas
                resumo["suporte_valor"] += item.valor

        for responsavel, resumo in sorted(agrupado.items()):
            horas_totais = resumo["sobre_horas"] + resumo["suporte_horas"]
            valor_total = resumo["sobre_valor"] + resumo["suporte_valor"]
            ws.append(
                [
                    responsavel,
                    round(resumo["sobre_horas"], 2),
                    round(resumo["sobre_valor"], 2),
                    round(resumo["suporte_horas"], 2),
                    round(resumo["suporte_valor"], 2),
                    round(horas_totais, 2),
                    round(valor_total, 2),
                ]
            )
        self._stylize_table(ws, currency_columns={3, 5, 7}, hours_columns={2, 4, 6})

    def _fill_summary_by_month_type(self, ws, projeto: ProjetoPlantao) -> None:
        header = ["Mês", "Tipo", "Qtd. Lançamentos", "Horas Totais", "Valor Total"]
        ws.append(header)

        agrupado: dict[tuple[str, str], dict] = {}
        for item in projeto.lancamentos:
            chave = (item.mes, item.tipo)
            resumo = agrupado.setdefault(chave, {"quantidade": 0, "horas": 0.0, "valor": 0.0})
            resumo["quantidade"] += 1
            resumo["horas"] += item.total_horas
            resumo["valor"] += item.valor

        ordem_meses = {
            "Janeiro": 1,
            "Fevereiro": 2,
            "Março": 3,
            "Abril": 4,
            "Maio": 5,
            "Junho": 6,
            "Julho": 7,
            "Agosto": 8,
            "Setembro": 9,
            "Outubro": 10,
            "Novembro": 11,
            "Dezembro": 12,
        }
        for (mes, tipo), resumo in sorted(
            agrupado.items(), key=lambda item: (ordem_meses.get(item[0][0], 99), item[0][1])
        ):
            ws.append(
                [
                    mes,
                    tipo,
                    resumo["quantidade"],
                    round(resumo["horas"], 2),
                    round(resumo["valor"], 2),
                ]
            )
        self._stylize_table(ws, currency_columns={5}, hours_columns={4})

    def _stylize_table(self, ws, currency_columns: set[int], hours_columns: set[int]) -> None:
        header_fill = PatternFill("solid", fgColor="1F5AA6")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin", color="D7E3F5"),
            right=Side(style="thin", color="D7E3F5"),
            top=Side(style="thin", color="D7E3F5"),
            bottom=Side(style="thin", color="D7E3F5"),
        )
        alignment = Alignment(vertical="center", horizontal="left")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = alignment
                cell.border = border
                if cell.column in currency_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
                if cell.column in hours_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

        for index, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                current = "" if cell.value is None else str(cell.value)
                if len(current) > max_length:
                    max_length = len(current)
            ws.column_dimensions[get_column_letter(index)].width = min(max_length + 4, 34)
